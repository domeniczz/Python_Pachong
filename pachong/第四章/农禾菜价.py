import os
import requests
from lxml import etree
from openpyxl import Workbook, load_workbook
from concurrent.futures import ThreadPoolExecutor


def writeExcelTable(table, ind):
    market_place = table.xpath("./p[1]/text()")[0].strip()
    each = table.xpath("./div[1]/table/tbody/tr")
    ws = wb.create_sheet(market_place, ind)

    # 获取列标题
    # ws['a1'] = table.xpath("./div[1]/table/thead/tr/th[1]/text()")[0].strip()  # "序号"
    # ws['b1'] = table.xpath("./div[1]/table/thead/tr/th[2]/text()")[0].strip()  # "菜名"
    # ws['c1'] = table.xpath("./div[1]/table/thead/tr/th[3]/text()")[0].strip()  # "单位"
    # ws['d1'] = table.xpath("./div[1]/table/thead/tr/th[4]/text()")[0].strip()  # "价格"
    # ws['e1'] = table.xpath("./div[1]/table/thead/tr/th[5]/text()")[0].strip()  # "升跌"
    # ws['f1'] = table.xpath("./div[1]/table/thead/tr/th[6]/text()")[0].strip()  # "升跌(%)"
    temp = table.xpath("./div[1]/table/thead/tr/th/text()")
    ws['a1'] = temp[0].strip()  # "序号"
    ws['b1'] = temp[1].strip()  # "菜名"
    ws['c1'] = temp[2].strip()  # "单位"
    ws['d1'] = temp[3].strip()  # "价格"
    ws['e1'] = temp[4].strip()  # "升跌"
    ws['f1'] = temp[5].strip()  # "升跌(%)"

    for e in each:
        index = e.xpath("./td[1]/text()")[0].strip()
        name = e.xpath("./td[2]/text()")[0].strip()
        unit = e.xpath("./td[3]/text()")[0].strip()
        price = e.xpath("./td[4]/text()")[0].strip()
        rise = e.xpath("./td[5]/span/text()")[0].strip()
        percent = e.xpath("./td[6]/span/text()")[0].strip()

        ws.append((index, name, unit, price, rise, percent))  # 用元组

    print("Write " + market_place + " Over" + f"\n{len(each)} records in total\n")


if __name__ == '__main__':
    url = "https://nonghecj.com/"

    os.environ['NO_PROXY'] = "https://nonghecj.com/"  # 意思是设置系统变量，['NO_PROXY']的意思是指定域名别用代理来处理

    headers = {
        "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/96.0.4664.93 Safari/537.36"
    }

    resp = requests.get(url, headers=headers, verify=False)
    resp.encoding = "UTF-8"
    print(resp, '\n')

    tree = etree.HTML(resp.text)
    blocks = tree.xpath("/html/body/div[1]/div/div/div[2]/div[2]/div[1]/div[2]")[0]
    tables = blocks.xpath("./div")

    # 文件的安全打开：如果有文件就打开，如果没有就新建
    if os.path.exists("农禾菜价.xlsx"):
        wb = load_workbook("农禾菜价.xlsx")

        # 清空所有sheet
        all_sheets = wb.sheetnames  # 得到所有sheet的名称
        for i in range(len(all_sheets)):
            sheet = wb[all_sheets[i]]  # 通过名称得到对应的sheet
            wb.remove(sheet)  # 删除sheet
    else:
        wb = Workbook("农禾菜价.xlsx")

    # 多线程（使用线程池），线程数为 round(len(tables)/3)
    with ThreadPoolExecutor(round(len(tables)/3)) as t:
        i = 0  # 记录第几个market_place
        for ts in tables:
            writeExcelTable(ts, i)
            i += 1
    print("Threads Over!")

    # i = 0  # 记录第几个market_place
    # for t in tables:
    #     writeExcelTable(t, i)
    #     i += 1

    resp.close()
    wb.save("农禾菜价.xlsx")
    wb.close()
